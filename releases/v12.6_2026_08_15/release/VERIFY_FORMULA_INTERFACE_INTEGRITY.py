#!/usr/bin/env python3
"""Verify the v5.6 workbook formula and qualification-continuity interface."""
from __future__ import annotations
from pathlib import Path
import hashlib, sys
from openpyxl import load_workbook

sys.dont_write_bytecode = True
ROOT = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else Path(__file__).resolve().parents[1]
BOOK = ROOT / "docs/aligners/RippleLogic_Aligners_Sheet_v5.6.xlsx"
MIRROR = ROOT / "core_15/RippleLogic_Aligners_Sheet_v5.6.xlsx"

if hashlib.sha256(BOOK.read_bytes()).digest() != hashlib.sha256(MIRROR.read_bytes()).digest():
    raise SystemExit("FAIL workbook interface: Core 15 mirror mismatch")
formulas = load_workbook(BOOK, data_only=False, read_only=True)
values = load_workbook(BOOK, data_only=True, read_only=True)
required = {"Qualification_Continuity", "Outcome_Requalification", "v12_6_Sync"}
if len(formulas.sheetnames) != 87 or not required.issubset(formulas.sheetnames):
    raise SystemExit("FAIL workbook interface: sheet inventory")

formula_count = 0
errors = []
for sheet in formulas.sheetnames:
    for row in formulas[sheet].iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1
                cached = values[sheet][cell.coordinate].value
                if isinstance(cached, str) and cached.startswith("#"):
                    errors.append((sheet, cell.coordinate, cached))
if formula_count != 1643 or errors:
    raise SystemExit(f"FAIL workbook interface: formulas={formula_count} cached_errors={errors[:5]}")

checks = {
    ("Qualification_Continuity", "F5"): ("=IF(", "MATCH"),
    ("Qualification_Continuity", "F28"): ("BLOCK_OR_REQUALIFY", "QUALIFIED_RECORD_COMPLETE"),
}
for (sheet, cell), (formula_token, cached_value) in checks.items():
    formula = formulas[sheet][cell].value
    cached = values[sheet][cell].value
    if formula_token not in str(formula) or cached != cached_value:
        raise SystemExit(f"FAIL workbook interface: {sheet}!{cell} formula/cache")
sync = " ".join(str(values["v12_6_Sync"][f"B{row}"].value) for row in range(2, 28))
for token in ("v12.6", "v8.5", "v5.6", "v1.6", "v4"):
    if token not in sync:
        raise SystemExit(f"FAIL workbook interface: sync token {token}")
print(f"PASS workbook v5.6 interface: 87 sheets, {formula_count} formulas, zero cached errors, exact action binding, and exact Core 15 mirror")
