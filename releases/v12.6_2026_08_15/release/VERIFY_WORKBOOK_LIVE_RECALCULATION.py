#!/usr/bin/env python3
"""Hard-recalculate the v5.6 workbook in an isolated LibreOffice profile."""
from __future__ import annotations

import os
import shutil
import subprocess
import sys
import tempfile
from itertools import zip_longest
from pathlib import Path

from openpyxl import load_workbook

sys.dont_write_bytecode = True
ROOT = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else Path(__file__).resolve().parents[1]
SOURCE = ROOT / "docs/aligners/RippleLogic_Aligners_Sheet_v5.6.xlsx"

with tempfile.TemporaryDirectory(prefix="mathgov_v56_recalc_") as raw:
    tmp = Path(raw)
    profile = tmp / "profile"
    profile.mkdir()
    out = tmp / "out"
    out.mkdir()
    candidate = tmp / "input.xlsx"
    shutil.copy2(SOURCE, candidate)
    env = dict(os.environ)
    env["PYTHONDONTWRITEBYTECODE"] = "1"
    result = subprocess.run(
        [
            "soffice",
            f"-env:UserInstallation=file://{profile}",
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(out),
            str(candidate),
        ],
        text=True,
        capture_output=True,
        env=env,
        timeout=120,
    )
    recalculated = out / "input.xlsx"
    if result.returncode or not recalculated.is_file():
        raise SystemExit(f"FAIL workbook live recalculation: {result.stdout}{result.stderr}")

    formulas = load_workbook(recalculated, data_only=False, read_only=True)
    values = load_workbook(recalculated, data_only=True, read_only=True)
    if formulas.sheetnames != values.sheetnames:
        raise SystemExit("FAIL workbook live recalculation: formula/value sheet inventory mismatch")

    count = 0
    errors: list[tuple[str, str, str]] = []
    for sheet in formulas.sheetnames:
        fws = formulas[sheet]
        vws = values[sheet]
        for frow, vrow in zip_longest(fws.iter_rows(), vws.iter_rows(), fillvalue=()):
            for fcell, vcell in zip_longest(frow, vrow, fillvalue=None):
                if fcell is None:
                    continue
                if isinstance(fcell.value, str) and fcell.value.startswith("="):
                    count += 1
                    value = None if vcell is None else vcell.value
                    if isinstance(value, str) and value.startswith("#"):
                        errors.append((sheet, fcell.coordinate, value))

    if count != 1643 or errors:
        raise SystemExit(f"FAIL workbook live recalculation: formulas={count} errors={errors[:5]}")
    if values["Qualification_Continuity"]["F28"].value != "QUALIFIED_RECORD_COMPLETE":
        raise SystemExit("FAIL workbook live recalculation: qualification continuity result")

print("PASS isolated LibreOffice recalculation: 1,643 formulas, zero formula errors, qualification record complete")
